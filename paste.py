import cv2
import math
import os
import pathlib
import numpy as np
from PIL import Image

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as xlImage
from openpyxl.styles.borders import Border, Side

def imread(file_name, flags=cv2.IMREAD_COLOR, dtype=np.uint8):
    """ imread for dealing with path including japanese """
    try:
        np_array = np.fromfile(file_name, dtype)
        img = cv2.imdecode(np_array, flags)
        return img
    except Exception as exc:
        print(exc)
        return None


def imwrite(file_name, img, params=None):
    """ imwrite for dealing with path including japanese """
    try:
        ext = os.path.splitext(file_name)[1]
        result, np_array = cv2.imencode(ext, img, params)

        if result:
            with open(file_name, mode='w+b') as file:
                np_array.tofile(file)
            return True
        else:
            return False

    except Exception as exc:
        print(exc)
        return False

def cv2pil(img):
    """ convert cv2 image to PIL image """
    tmp = img.copy()
    if tmp.ndim == 2:
        pass
    elif tmp.shape[2] == 3:
        tmp = cv2.cvtColor(tmp, cv2.COLOR_BGR2RGB)
    elif tmp.shape[2] == 4:
        tmp = cv2.cvtColor(tmp, cv2.COLOR_BGRA2RGBA)
    pil_img = Image.fromarray(tmp)
    return pil_img


def get_cell_index_letter(row, col):
    """ return index for Excel cell """
    return get_column_letter(col) + str(row)


def get_dir_list(root_path):
    """ return all directories in root dir """
    path_tmp = pathlib.Path(root_path)
    dirs = [str(path) for path in path_tmp.glob('**') if path.is_dir()]
    return dirs

class App:
    """ Paste images in directories to worksheet """

    def __init__(self):
        self.set_info()


    def set_info(self):
        """ initialize """
        self.cell_width_pix = 88
        self.cell_height_pix = 18

        self.img_insert_width = self.cell_width_pix * 3
        self.cols_per_img = 4

        self.max_cols = 200
        self.max_rows = 1000


    def format_cells(self, worksheet):
        """ resize cells in worksheet and draw lines """
        if self.cell_width_pix != 88:
            for col in range(1, self.max_cols):
                worksheet.column_dimensions[get_column_letter(col)].width = self.cell_width_pix / 8
        if self.cell_height_pix != 18:
            for row in range(1, self.max_rows):
                worksheet.row_dimensions[row].height = self.cell_height_pix / 1.3


    def img_resize(self, img):
        """ resize image """
        self.image_height, self.image_width = img.shape[:2]
        size = (self.img_insert_width, self.image_height * self.img_insert_width // self.image_width)
        resized_img = cv2.resize(img, size)
        self.image_width = size[0]
        self.image_height = size[1]

        self.rows_per_img = math.ceil(self.image_height / self.cell_height_pix)

        return resized_img


    def write_info(self, worksheet, row, col, info, prefix_str = ''):
        """
        write info on cell(row, col) in worksheet
        
        [example] write_info(ws, 'C:/hoge/huga.png', 2, 2, 'Path: ')
            written as "Path: C:/hoge/huga.png" on cell(2, 2) in ws
        """
        write_idx_letter = get_cell_index_letter(row, col)
        worksheet[write_idx_letter] = prefix_str + info


    def get_next_col(self, crr_col):
        """ return next cell column index for inserting image """
        return crr_col + self.cols_per_img

    def get_next_row(self, crr_row):
        """ return next cell row index for inserting image """
        return crr_row + self.rows_per_img + 3


    def execute(self, worksheet, root_dir, image_names):
        """ main function """
        self.max_cols = 1 + self.cols_per_img * len(image_names)

        border_top = Border(top=Side(style='thick', color='000000'))
        border_right_top = Border(right=Side(style='thick', color='000000'), top=Side(style='thick', color='000000'))
        border_right = Border(right=Side(style='thick', color='000000'))

        self.format_cells(worksheet)

        pic_insert_row = 2
        pic_insert_col = 2

        dirs = get_dir_list(root_dir)

        max_dir_len = 0
        for directory in dirs:
            # image exist check
            path_tmp = pathlib.Path(directory)
            files = [path for path in path_tmp.glob('./*') if path.is_file()]
            if len(files) == 0:
                continue

            for col in range(1, self.max_cols + 1):
                cell_idx_letter = get_cell_index_letter(pic_insert_row, col)
                worksheet[cell_idx_letter].border = border_top
            self.write_info(worksheet, pic_insert_row, 1, directory)

            max_dir_len = max(max_dir_len, len(directory))

            for img_name in image_names:

                img_path = directory + '\\' + img_name
                
                # write path to cell
                self.write_info(worksheet, pic_insert_row, pic_insert_col + 1, img_name)

                # load image
                tmp_img = imread(img_path)
                if not tmp_img is None:
                    tmp_img = self.img_resize(tmp_img)

                    # write image to cell
                    pic_insert_idx_letter = get_cell_index_letter(pic_insert_row + 1, pic_insert_col + 1)
                    img = xlImage(cv2pil(tmp_img))
                    worksheet.add_image(img, pic_insert_idx_letter)

                pic_insert_col = self.get_next_col(pic_insert_col)

            for dr in range(self.rows_per_img + 3):
                cell_idx_letter = get_cell_index_letter(pic_insert_row + dr, 1)
                if dr == 0:
                    worksheet[cell_idx_letter].border = border_right_top
                    continue
                worksheet[cell_idx_letter].border = border_right

            pic_insert_col = 2
            pic_insert_row = self.get_next_row(pic_insert_row)

        worksheet.column_dimensions[get_column_letter(1)].width = max_dir_len


def main():
    """ main """

    out_xlsx_path = 'test.xlsx'
    root_dir = 'test_dir'
    image_names = ['test.bmp', 'red.bmp', 'green.bmp', 'blue.bmp']

    workbook = openpyxl.Workbook()
    sheet_names = workbook.get_sheet_names()
    worksheet = workbook[sheet_names[0]]

    app = App()
    app.execute(worksheet, root_dir, image_names)

    workbook.save(out_xlsx_path)


if __name__ == '__main__':
    main()
